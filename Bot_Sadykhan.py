import logging
import os
import csv
import pytz
from datetime import datetime
from dotenv import load_dotenv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import asyncio
import json
from logging.handlers import RotatingFileHandler

from aiogram import Bot, Dispatcher, F, types
from aiogram.enums import ParseMode
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import FSInputFile, Update
from aiogram.client.default import DefaultBotProperties
from aiohttp import web

# === Настройка логирования ===
logger = logging.getLogger("BotSadykhan")
logger.setLevel(logging.DEBUG)

formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")

console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

file_handler = RotatingFileHandler("app.log", maxBytes=5*1024*1024, backupCount=3, encoding="utf-8")
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# === Загрузка конфигов ===
load_dotenv()
API_TOKEN = os.getenv("API_TOKEN")
CHAT_ID = int(os.getenv("CHAT_ID", "0"))
TEMPLATE_PATH = os.getenv("TEMPLATE_PATH", "template.xlsx")
CHECKLIST_PATH = os.getenv("CHECKLIST_PATH", "checklist.xlsx")
LOG_PATH = os.getenv("LOG_PATH", "checklist_log.csv")
PORT = int(os.getenv("PORT", 8080))
WEBHOOK_URL = os.getenv("WEBHOOK_URL")

logger.info("Bot configuration loaded.")
logger.debug("API_TOKEN is set: [REDACTED]")
logger.debug(f"CHAT_ID: {CHAT_ID}")
logger.debug(f"TEMPLATE_PATH: {TEMPLATE_PATH}")
logger.debug(f"CHECKLIST_PATH: {CHECKLIST_PATH}")
logger.debug(f"LOG_PATH: {LOG_PATH}")
logger.debug(f"PORT: {PORT}")
logger.debug(f"WEBHOOK_URL: {WEBHOOK_URL}")

# === FSM-состояния ===
class Form(StatesGroup):
    name = State()
    pharmacy = State()
    rating = State()
    comment = State()

# === Читаем критерии из Excel ===
criteria = []
try:
    logger.info(f"Reading checklist from: {CHECKLIST_PATH}")
    df = pd.read_excel(CHECKLIST_PATH, sheet_name="Чек лист", header=None)
    start_i = df[df.iloc[:, 0] == "Блок"].index[0] + 1
    df = df.iloc[start_i:, :8].reset_index(drop=True)
    df.columns = ["Блок", "Критерий", "Требование", "Оценка", "Макс", "Примечание", "Дата проверки", "Дата исправления"]

    last_block = "Неизвестный блок"
    for _, row in df.iterrows():
        if pd.isna(row["Критерий"]) or pd.isna(row["Требование"]):
            continue

        block = row["Блок"] if pd.notna(row["Блок"]) else last_block
        last_block = block
        max_value = str(row["Макс"])
        max_score = int(max_value) if max_value.isdigit() else 10
        criteria.append({
            "block": block,
            "criterion": str(row["Критерий"]),
            "requirement": str(row["Требование"]),
            "max": max_score
        })
    logger.info(f"Loaded {len(criteria)} criteria.")
    logger.debug(f"Criteria content: {criteria}")
except Exception as e:
    logger.error(f"Error reading checklist: {e}", exc_info=True)

# === Утилиты ===
def now_ts():
    return datetime.now(pytz.timezone("Asia/Almaty")).strftime("%Y-%m-%d %H:%M:%S")

def log_csv(pharm, name, ts, score, total):
    first = not os.path.exists(LOG_PATH)
    try:
        with open(LOG_PATH, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if first:
                writer.writerow(["Дата", "Аптека", "Проверяющий", "Баллы", "Макс"])
            writer.writerow([ts, pharm, name, score, total])
    except Exception as e:
        logger.error(f"Error writing to log: {e}", exc_info=True)

# === Инициализация бота ===
bot = Bot(token=API_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())

# === Команда /start ===
async def cmd_start(msg: types.Message, state: FSMContext):
    logger.info(f"User {msg.from_user.id} called /start")
    await state.clear()
    await msg.answer(
        "<b>📋 Чек‑лист посещения аптек</b>\n\n"
        "Введите ваше ФИО для авторизации:",
        parse_mode=ParseMode.HTML
    )
    await state.set_state(Form.name)

# === /id для отладки ===
async def cmd_id(msg: types.Message):
    logger.info(f"User {msg.from_user.id} called /id")
    await msg.answer(f"<code>{msg.chat.id}</code>")

# === Сброс FSM ===
async def cmd_reset(msg: types.Message, state: FSMContext):
    logger.info(f"User {msg.from_user.id} called /сброс")
    await state.clear()
    await msg.answer("Состояние сброшено. /start — начать заново")

# === Обработка ФИО ===
async def proc_name(msg: types.Message, state: FSMContext):
    name = msg.text.strip()
    logger.info(f"User {msg.from_user.id} entered name: {name}")
    await state.update_data(name=name, step=0, data=[], start=now_ts())
    await msg.answer("Введите название аптеки:")
    await state.set_state(Form.pharmacy)

# === Обработка названия аптеки ===
async def proc_pharmacy(msg: types.Message, state: FSMContext):
    pharmacy = msg.text.strip()
    logger.info(f"User {msg.from_user.id} entered pharmacy: {pharmacy}")
    await state.update_data(pharmacy=pharmacy)
    await msg.answer("Начинаем проверку…")
    await state.set_state(Form.rating)
    logger.debug(f"Calling send_question for chat {msg.chat.id}")
    await send_question(msg.chat.id, state)

# === Отправка вопроса ===
async def send_question(chat_id: int, state: FSMContext):
    logger.info(f"Sending question to chat {chat_id}")
    data = await state.get_data()
    step = data.get("step", 0)
    total = len(criteria)
    logger.debug(f"Step: {step}, Total: {total}")

    if total == 0:
        logger.error("Criteria list is empty")
        await bot.send_message(chat_id, "❌ Ошибка: Не удалось загрузить критерии проверки.")
        await state.clear()
        return

    if step >= total:
        logger.info(f"All criteria processed for chat {chat_id}")
        await bot.send_message(
            chat_id,
            "✅ Все оценки поставлены!\n\n"
            "📝 Напишите ваши выводы по аптеке:",
            parse_mode=ParseMode.HTML
        )
        await state.set_state(Form.comment)
        return

    try:
        criterion = criteria[step]
        logger.debug(f"Criterion at step {step}: {criterion}")
        required_keys = ["block", "criterion", "requirement", "max"]
        missing_keys = [key for key in required_keys if key not in criterion or criterion[key] is None or str(criterion[key]).strip() == ""]
        if missing_keys:
            logger.error(f"Missing or empty keys in criterion at step {step}: {missing_keys}")
            await bot.send_message(chat_id, f"❌ Ошибка: Некорректные данные критерия на шаге {step + 1}.")
            await state.clear()
            return

        text = (
            f"<b>Вопрос {step + 1} из {total}</b>\n\n"
            f"<b>Блок:</b> {criterion['block']}\n"
            f"<b>Критерий:</b> {criterion['criterion']}\n"
            f"<b>Требование:</b> {criterion['requirement']}\n"
            f"<b>Макс. балл:</b> {criterion['max']}"
        )
        kb = InlineKeyboardBuilder()
        start_score = 0 if criterion["max"] == 1 else 1
        for i in range(start_score, criterion["max"] + 1):
            kb.button(text=str(i), callback_data=f"score_{i}")
        if step > 0:
            kb.button(text="◀️ Назад", callback_data="prev")
        kb.adjust(5)

        sent_message = await bot.send_message(chat_id, text, reply_markup=kb.as_markup())
        logger.debug(f"Sent question {step + 1} to chat {chat_id}, message_id: {sent_message.message_id}")
    except Exception as e:
        logger.error(f"Error in send_question for chat {chat_id}: {e}", exc_info=True)
        await bot.send_message(chat_id, "❌ Ошибка при отправке вопроса. Пожалуйста, начните заново с /start.")
        await state.clear()

# === Обработка callback-запросов ===
async def cb_all(cb: types.CallbackQuery, state: FSMContext):
    logger.info(f"Callback from user {cb.from_user.id}: {cb.data}")
    await cb.answer()

    data = await state.get_data()
    logger.debug(f"FSM state data: {data}")

    if not data or "step" not in data:
        logger.error("FSM state is empty or missing 'step' key")
        await bot.send_message(cb.message.chat.id, "❌ Ошибка: Состояние сброшено. Пожалуйста, начните заново с /start.")
        await state.clear()
        return

    step = data.get("step", 0)
    total = len(criteria)
    logger.debug(f"Step: {step}, Total: {total}, Data: {cb.data}")

    if step >= total:
        logger.debug("All criteria rated")
        return

    if cb.data.startswith("score_"):
        try:
            score = int(cb.data.split("_")[1])
            criterion = criteria[step]
            if score <= criterion["max"]:
                # Проверяем наличие ключа 'data'
                if "data" not in data or not isinstance(data["data"], list):
                    logger.warning(f"'data' key missing or invalid in FSM state, initializing as empty list")
                    data["data"] = []
                    await bot.send_message(
                        cb.message.chat.id,
                        "⚠️ Ошибка: Данные были потеряны. Пожалуйста, начните проверку заново с /start."
                    )
                    await state.clear()
                    return
                data["data"].append({"crit": criterion, "score": score})
                data["step"] = step + 1
                await state.update_data(**data)
                logger.debug(f"Score {score} saved for step {step}")
                logger.debug(f"Total scores saved: {len(data['data'])}")
                # Проверяем, что количество сохранённых оценок соответствует текущему шагу
                expected_scores = step + 1
                if len(data["data"]) != expected_scores:
                    logger.error(f"Data inconsistency: expected {expected_scores} scores, but found {len(data['data'])}")
                    await bot.send_message(
                        cb.message.chat.id,
                        f"❌ Ошибка: Данные повреждены (ожидается {expected_scores} оценок, найдено {len(data['data'])}). Начните заново с /start."
                    )
                    await state.clear()
                    return
                try:
                    await bot.edit_message_text(
                        f"✅ Оценка: {score} {'⭐' * score}",
                        chat_id=cb.message.chat.id,
                        message_id=cb.message.message_id
                    )
                except Exception as e:
                    logger.error(f"Error editing message: {e}", exc_info=True)
                await send_question(cb.message.chat.id, state)
            else:
                logger.warning(f"Invalid score {score} for max {criterion['max']}")
                await bot.send_message(cb.message.chat.id, "❌ Неверная оценка, попробуйте снова.")
        except ValueError as e:
            logger.error(f"Invalid callback data: {cb.data}, error: {e}")
            await bot.send_message(cb.message.chat.id, "❌ Ошибка обработки оценки.")
    elif cb.data == "prev" and step > 0:
        if "data" not in data or not isinstance(data["data"], list):
            logger.warning(f"'data' key missing or invalid in FSM state, initializing as empty list")
            data["data"] = []
            await bot.send_message(
                cb.message.chat.id,
                "⚠️ Ошибка: Данные были потеряны. Пожалуйста, начните проверку заново с /start."
            )
            await state.clear()
            return
        data["step"] = step - 1
        data["data"].pop()
        await state.update_data(**data)
        logger.debug(f"Navigated back to step {step - 1}")
        logger.debug(f"Total scores saved: {len(data['data'])}")
        # Проверяем количество сохранённых оценок после возврата назад
        expected_scores = step
        if len(data["data"]) != expected_scores:
            logger.error(f"Data inconsistency after prev: expected {expected_scores} scores, but found {len(data['data'])}")
            await bot.send_message(
                cb.message.chat.id,
                f"❌ Ошибка: Данные повреждены (ожидается {expected_scores} оценок, найдено {len(data['data'])}). Начните заново с /start."
            )
            await state.clear()
            return
        await send_question(cb.message.chat.id, state)
    else:
        logger.warning(f"Unhandled callback: {cb.data}")
        await bot.send_message(cb.message.chat.id, "❌ Неизвестная команда.")

# === Обработка комментария ===
async def proc_comment(msg: types.Message, state: FSMContext):
    comment = msg.text.strip()
    logger.info(f"User {msg.from_user.id} entered comment: {comment}")
    data = await state.get_data()
    data["comment"] = comment
    await state.update_data(**data)
    logger.debug(f"Data before report generation: {data}")
    if "data" not in data or not data["data"]:
        logger.warning("No scores saved for the report")
        await msg.answer("⚠️ Ошибка: Оценки не сохранены. Пожалуйста, начните проверку заново с /start.")
        await state.clear()
        return
    # Проверяем, что все 32 шага завершены
    total_steps = len(criteria)
    saved_scores = len(data["data"])
    logger.info(f"Total scores saved before report: {saved_scores}")
    if saved_scores != total_steps:
        logger.error(f"Expected {total_steps} scores, but found {saved_scores}")
        await msg.answer(
            f"❌ Ошибка: Завершено только {saved_scores} из {total_steps} шагов. Начните проверку заново с /start."
        )
        await state.clear()
        return
    await msg.answer("⌛ Формирую отчёт…")
    await make_report(msg.chat.id, data)
    await state.clear()

# === Генерация отчёта ===
async def make_report(user_id: int, data):
    logger.info(f"Generating report for user {user_id}")
    logger.debug(f"Report data: {data}")
    name = data["name"]
    ts = data["start"]
    pharmacy = data["pharmacy"]
    report_filename = f"{pharmacy}_{name}_{datetime.strptime(ts, '%Y-%m-%d %H:%M:%S').strftime('%d.%m.%Y_%H%M')}.xlsx".replace(" ", "_")

    try:
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        title = f"Отчёт по проверке аптеки\nИсполнитель: {name}\nДата и время: {ts}"
        ws.merge_cells("A1:G2")
        ws["A1"] = title
        ws["A1"].font = Font(size=14, bold=True)
        ws["B3"] = pharmacy

        headers = ["Блок", "Критерий", "Требование", "Оценка", "Макс", "Коммент.", "Дата проверки"]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(5, idx, header)
            cell.font = Font(bold=True)

        row = 6
        total_score = 0
        total_max = 0
        if "data" in data and data["data"]:
            processed_count = 0
            for rec in data["data"]:
                crit = rec["crit"]
                score = rec["score"]
                ws.cell(row, 1, crit["block"])
                ws.cell(row, 2, crit["criterion"])
                ws.cell(row, 3, crit["requirement"])
                ws.cell(row, 4, score)
                ws.cell(row, 5, crit["max"])
                ws.cell(row, 6, "")
                ws.cell(row, 7, ts)
                total_score += score
                total_max += crit["max"]
                row += 1
                processed_count += 1
            logger.info(f"Processed {processed_count} records in report")
        else:
            logger.warning("No data available for report, table will be empty")
            await bot.send_message(user_id, "⚠️ Внимание: Отчёт пуст, так как оценки не были сохранены.")

        ws.cell(row + 1, 3, "ИТОГО:")
        ws.cell(row + 1, 4, total_score)
        ws.cell(row + 2, 3, "Максимум:")
        ws.cell(row + 2, 4, total_max)
        ws.cell(row + 4, 1, "Вывод проверяющего:")
        ws.cell(row + 5, 1, data.get("comment", ""))

        wb.save(report_filename)
        logger.info(f"Report saved: {report_filename}")

        try:
            file = FSInputFile(report_filename, filename=report_filename)
            await bot.send_document(user_id, file)
            logger.info(f"Report sent to user {user_id}")
        except Exception as e:
            logger.error(f"Error sending report: {e}", exc_info=True)
            await bot.send_message(user_id, "❌ Ошибка при отправке отчёта.")
            return

        log_csv(pharmacy, name, ts, total_score, total_max)
        await bot.send_message(user_id, "✅ Отчёт сформирован и отправлен.\n/start — новая проверка")

    except Exception as e:
        logger.error(f"Error generating report: {e}", exc_info=True)
        await bot.send_message(user_id, "❌ Ошибка при формировании отчёта.")
    finally:
        try:
            if os.path.exists(report_filename):
                os.remove(report_filename)
                logger.info(f"Deleted report file: {report_filename}")
        except Exception as e:
            logger.warning(f"Failed to delete report file: {e}")

# === Webhook ===
async def handle_webhook(request: web.Request):
    logger.info(f"Webhook received: {request.method} {request.url}")
    logger.debug(f"Request headers: {request.headers}")
    try:
        update = await request.json()
        logger.debug(f"Webhook data: {json.dumps(update, indent=2, ensure_ascii=False)}")
        update_obj = Update(**update)
        await dp.feed_update(bot, update_obj)
        logger.info("Webhook update processed successfully")
        return web.Response(text="OK")
    except Exception as e:
        logger.error(f"Webhook error: {e}", exc_info=True)
        return web.Response(status=500)

async def on_startup(bot: Bot):
    if WEBHOOK_URL:
        webhook_path = "/webhook"
        webhook_url = f"{WEBHOOK_URL}{webhook_path}"
        logger.info(f"Attempting to set webhook to: {webhook_url}")
        try:
            current_webhook = await bot.get_webhook_info()
            logger.debug(f"Current webhook info: {current_webhook}")
            await bot.set_webhook(webhook_url)
            logger.info(f"Webhook successfully set to: {webhook_url}")
            updated_webhook = await bot.get_webhook_info()
            logger.debug(f"Updated webhook info: {updated_webhook}")
        except Exception as e:
            logger.error(f"Error setting webhook: {e}", exc_info=True)
            logger.warning("Falling back to long polling due to webhook failure")
            return False
    else:
        logger.warning("WEBHOOK_URL not set, using long polling")
        return False
    return True

async def on_shutdown(bot: Bot):
    logger.info("Shutting down bot")
    await bot.delete_webhook()
    await bot.session.close()

async def main():
    dp.startup.register(on_startup)
    dp.shutdown.register(on_shutdown)

    dp.message.register(cmd_start, F.text == "/start")
    dp.message.register(cmd_id, F.text == "/id")
    dp.message.register(cmd_reset, F.text == "/сброс")
    dp.message.register(proc_name, Form.name)
    dp.message.register(proc_pharmacy, Form.pharmacy)
    dp.message.register(proc_comment, Form.comment)
    dp.callback_query.register(cb_all)

    use_webhook = await on_startup(bot)

    if use_webhook:
        app = web.Application()
        app.add_routes([web.post("/webhook", handle_webhook)])
        runner = web.AppRunner(app)
        await runner.setup()
        site = web.TCPSite(runner, "0.0.0.0", PORT)
        await site.start()
        logger.info(f"Webhook server started on port {PORT}")
        while True:
            await asyncio.sleep(3600)
    else:
        logger.info("Starting bot in long polling mode")
        await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
